import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- SETUP FILE PATHS ---
folder = "H:/report_gen"  # Your folder path
aps_path = os.path.join(folder, "aps.xlsx")
cio_path = os.path.join(folder, "cio.xlsx")
lookup_path = os.path.join(folder, "s.xlsx")
output_path = os.path.join(folder, "attestation_report.xlsx")

# --- LOAD DATA ---
print("Loading Excel files...")
aps_df = pd.read_excel(aps_path)
cio_df = pd.read_excel(cio_path)
lookup1 = pd.read_excel(lookup_path, sheet_name="Sheet1")    # Changed to Sheet1
lookup2 = pd.read_excel(lookup_path, sheet_name="Sheet2")    # Changed to Sheet2
print(f"aps_df shape: {aps_df.shape}")
print(f"cio_df shape: {cio_df.shape}")
print(f"lookup1 columns: {lookup1.columns.tolist()}")
print(f"lookup2 columns: {lookup2.columns.tolist()}")

# --- CLEAN & RENAME APS FILE ---
print("Cleaning and renaming APS file columns...")
try:
    aps_df.columns = ['contextid', 'context name', 'title', 'status', 'due date', 'completion date', 'accountable']
except Exception as e:
    print(f"Error renaming aps_df columns: {e}")
print(f"aps_df columns after rename: {aps_df.columns.tolist()}")

aps_df = aps_df.drop(columns=aps_df.columns[7:], errors='ignore')  # Delete H-J if exist
aps_df.rename(columns={
    'completion date': 'Application Owner',
    'accountable': 'APS Accountable'
}, inplace=True)
print(f"aps_df columns after drop & rename: {aps_df.columns.tolist()}")

# --- CLEAN & RENAME CIO FILE ---
print("Cleaning and renaming CIO file columns...")
try:
    cio_df.columns = ['contextid', 'context name', 'title', 'status', 'due date', 'completion date', 'accountable']
except Exception as e:
    print(f"Error renaming cio_df columns: {e}")
print(f"cio_df columns after rename: {cio_df.columns.tolist()}")

cio_df = cio_df.drop(columns=['completion date'], errors='ignore')  # Delete F
cio_df = cio_df.drop(columns=cio_df.columns[6:], errors='ignore')   # Delete G-I if exist
cio_df.rename(columns={
    'accountable': 'Application Owner'
}, inplace=True)
print(f"cio_df columns after drop & rename: {cio_df.columns.tolist()}")

# --- ADD COLUMNS BEFORE & AFTER Due Date ---
print("Inserting new columns...")
for df, name in zip([aps_df, cio_df], ["APS", "CIO"]):
    df.insert(4, 'CPPM Review Status', '')    # New col before due date
    df.insert(5, 'PECM_Delegate', '')         # New col after due date
    print(f"{name} df shape after insert: {df.shape}")
    print(f"{name} df columns: {df.columns.tolist()}")

# --- FINALIZE COLUMNS NAMES ---
print("Finalizing columns names...")
try:
    aps_df.columns = [
        'AIT #', 'AIT Name', 'Assessment Name', 'Trident status', 'CPPM Review Status',
        'Due Date', 'PECM_Delegate', 'Application Owner', 'APS Accountable'
    ]
except Exception as e:
    print(f"Error finalizing aps_df columns: {e}")

try:
    cio_df.columns = [
        'AIT #', 'AIT Name', 'Assessment Name', 'Trident status', 'CPPM Review Status',
        'Due Date', 'PECM_Delegate', 'Application Owner'
    ]
except Exception as e:
    print(f"Error finalizing cio_df columns: {e}")

print(f"Final APS columns: {aps_df.columns.tolist()}")
print(f"Final CIO columns: {cio_df.columns.tolist()}")

# --- ADD EMPTY FINAL COLUMNS ---
print("Adding empty final columns...")
for col in ['Support Owner', 'APS SLT', 'Tech Exec', 'CIO Exec']:
    aps_df[col] = ''
for col in ['APS Accountable', 'Support Owner', 'APS SLT', 'Tech Exec', 'CIO Exec']:
    cio_df[col] = ''

print("Columns after adding empty columns:")
print(f"APS columns: {aps_df.columns.tolist()}")
print(f"CIO columns: {cio_df.columns.tolist()}")

# --- PERFORM VLOOKUPS (MERGE) ---
print("Performing lookups/merges...")
lookup1 = lookup1[['app_id', 'Support Owner', 'APS SLT']]
lookup2 = lookup2[['app_id', 'Tech Exec', 'CIO Exec']]

aps_df = aps_df.merge(lookup1, how='left', left_on='AIT #', right_on='app_id').drop(columns='app_id')
aps_df = aps_df.merge(lookup2, how='left', left_on='AIT #', right_on='app_id').drop(columns='app_id')

cio_df = cio_df.merge(lookup1, how='left', left_on='AIT #', right_on='app_id').drop(columns='app_id')
cio_df = cio_df.merge(lookup2, how='left', left_on='AIT #', right_on='app_id').drop(columns='app_id')

print(f"APS after merges shape: {aps_df.shape}")
print(f"CIO after merges shape: {cio_df.shape}")

# --- CREATE PIVOT TABLES ---
print("Creating pivot tables...")
aps_pivot = aps_df.pivot_table(
    index=['CIO Exec', 'Tech Exec'],
    columns='Trident status',
    values='AIT #',
    aggfunc='count'
).reset_index()

cio_pivot = cio_df.pivot_table(
    index=['CIO Exec', 'Tech Exec'],
    columns='Trident status',
    values='AIT #',
    aggfunc='count'
).reset_index()

# --- CREATE MODIFIED PIVOTS (CPPM Review Status) ---
aps_mod_pivot = aps_df.pivot_table(
    index=['CIO Exec', 'Tech Exec'],
    columns='CPPM Review Status',
    values='AIT #',
    aggfunc='count'
).reset_index()

cio_mod_pivot = cio_df.pivot_table(
    index=['CIO Exec', 'Tech Exec'],
    columns='CPPM Review Status',
    values='AIT #',
    aggfunc='count'
).reset_index()

print("Pivot tables created.")

# --- EXPORT ALL TO EXCEL ---
print("Exporting to Excel...")
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    aps_df.to_excel(writer, sheet_name="APS Attestations", index=False)
    cio_df.to_excel(writer, sheet_name="CIO Attestations", index=False)
    aps_pivot.to_excel(writer, sheet_name="Apivot", index=False)
    cio_pivot.to_excel(writer, sheet_name="Cpivot", index=False)
    aps_mod_pivot.to_excel(writer, sheet_name="APS Summary", index=False)
    cio_mod_pivot.to_excel(writer, sheet_name="CIO Summary", index=False)

# --- APPLY FORMATTING TO FINAL SUMMARY SHEETS ---
print("Applying formatting...")
wb = load_workbook(output_path)

green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")   # Light blue
white_font = Font(color="FFFFFF", size=16, bold=True)
header_align = Alignment(horizontal="center", vertical="center")
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

def format_sheet(ws, fill_color):
    # Insert 3 rows at top to make space for title and header
    ws.insert_rows(1, amount=3)

    # Title row merged & formatted
    ws['A4'] = "APS Attestation summary" if ws.title == "APS Summary" else "CIO Attestation summary"
    ws.merge_cells('A4:F4')
    ws['A4'].font = white_font
    ws['A4'].alignment = header_align
    ws['A4'].fill = fill_color

    # Set Due Date header and value in row 1
    ws['B1'] = "Due Date"
    ws['B1'].value = "6/20/2025"
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}1'].alignment = header_align

    # Change row 3 col A text and bold it
    ws['A3'] = "Count of AITs"
    ws['A3'].font = Font(bold=True)

    # Highlight specific rows (4,9,18,20,23,29) cols A-G with fill color
    for row in [4, 9, 18, 20, 23, 29]:
        for col in 'ABCDEFG':
            cell = ws[f'{col}{row}']
            cell.fill = fill_color

    # Add borders to A-G for rows 1 to 31
    for row in ws.iter_rows(min_row=1, max_row=31, min_col=1, max_col=7):
        for cell in row:
            cell.border = border

    # Bold row 51 A-G (if exists)
    for col in 'ABCDEFG':
        cell = ws[f'{col}51']
        cell.font = Font(bold=True)

    # Autofit column widths A-G based on max length content
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

format_sheet(wb["APS Summary"], blue_fill)
format_sheet(wb["CIO Summary"], green_fill)

# Save formatted workbook
wb.save(output_path)

print(f"Done! Your formatted Excel file is saved here: {output_path}")